import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import datetime

st.set_page_config(page_title="Buyer-Seller Dashboard", layout="wide")
st.title("📊 Buyer-Seller Dashboard")

uploaded_file = st.file_uploader("📤 Upload your Excel file", type=["xls", "xlsx", "xlsm"])

# --- Helpers ---
NUMERIC_COLUMNS = [
    "Price", "price sold", "Amt", "Buyer Amt", "Misc Exp", "Profit",
    "Total Receivable", "Sel Wt", "Terms", "Days Seller", "Days Buyer"
]

def arrow_safe(df: pd.DataFrame) -> pd.DataFrame:
    # Ensure consistent types for Streamlit's Arrow serialization
    df = df.copy()
    # Normalize numeric columns to numeric (strings become NaN)
    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    # Convert remaining object columns to strings to avoid mixed-type issues
    for col in df.columns:
        if df[col].dtype == "object" and col not in ["Date"]:
            df[col] = df[col].astype(str)
    return df

# Load only the 'DATA' sheet
def load_data(uploaded_file):
    try:
        wb = load_workbook(uploaded_file, data_only=True, keep_vba=True)
        if "DATA" not in wb.sheetnames:
            st.error("❌ Sheet named 'DATA' not found.")
            return pd.DataFrame()
        sheet = wb["DATA"]
        data = list(sheet.values)
        headers = data[0]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=headers)

        # Format the 'Date' column to dd-mm-yyyy
        if "Date" in df.columns:
            date_series = pd.to_datetime(df["Date"], errors="coerce")
            df["Date"] = date_series.dt.strftime("%d-%m-%Y").fillna("")

        return df
    except Exception as e:
        st.error(f"❌ Error loading file: {e}")
        return pd.DataFrame()

# Export buttons
def export_buttons(df):
    st.markdown("### 📤 Export Options")
    file_name = st.text_input("📝 Enter export file name (without extension)", value="buyer_seller")

    # Use the same Arrow-safe df for export to avoid type surprises
    safe_df = arrow_safe(df)

    col1, col2 = st.columns(2)
    with col1:
        csv = safe_df.to_csv(index=False).encode('utf-8')
        st.download_button("📥 Export CSV", csv, f"{file_name}.csv", "text/csv")
    with col2:
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
            safe_df.to_excel(writer, index=False, sheet_name='DATA')
        st.download_button("📥 Export Excel", excel_buffer.getvalue(), f"{file_name}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Segmented views
def show_segmented_view(full_df):
    view = st.radio("📂 View Mode", ["All", "Buyer", "Seller"], horizontal=True)

    if view == "Buyer":
        df = full_df[full_df["Buyer"].notna()].copy()
    elif view == "Seller":
        df = full_df[full_df["Seller"].notna()].copy()
        df = df.iloc[:, :14]
    else:
        df = full_df.copy()

    if "Profit" in df.columns:
        df = df.drop(columns=["Profit"])

    filtered_df = apply_filters(df, full_df)

    if view == "Buyer":
        # Drop Id variants so only "Id Buyer" remains
        drop_cols = ["Id", "ID", "Price", "Terms", "Days Seller", "Amt", "Seller", "Broker"]
        filtered_df = filtered_df.drop(columns=[col for col in drop_cols if col in filtered_df.columns])
    elif view == "Seller":
        drop_cols = ["Id Buyer"]
        filtered_df = filtered_df.drop(columns=[col for col in drop_cols if col in filtered_df.columns])

    st.markdown(f"### 📋 Showing: {view} View")
    st.dataframe(arrow_safe(filtered_df), use_container_width=True)
    export_buttons(filtered_df)

# Advanced filters
def apply_filters(df, full_df):
    st.markdown("### 🔍 Advanced Filters")

    col1, col2 = st.columns(2)
    with col1:
        shape = st.multiselect("Shape", full_df["Shape"].dropna().unique()) if "Shape" in full_df.columns else []
        color = st.multiselect("Color", full_df["Color"].dropna().unique()) if "Color" in full_df.columns else []
        quality = st.multiselect("Quality", full_df["Quality"].dropna().unique()) if "Quality" in full_df.columns else []
        seller = st.multiselect("Seller", full_df["Seller"].dropna().unique()) if "Seller" in full_df.columns else []
    with col2:
        buyer = st.multiselect("Buyer", full_df["Buyer"].dropna().unique()) if "Buyer" in full_df.columns else []
        pointer = st.multiselect("Pointer", full_df["Pointer"].dropna().unique()) if "Pointer" in full_df.columns else []
        size = st.multiselect("Size (mm)", full_df["Size (mm)"].dropna().unique()) if "Size (mm)" in full_df.columns else []

        # Date range filter with default ignored
        ignore_date = st.checkbox("Ignore date filter (show all dates)", value=True)
        today = datetime.date.today()
        default_start = today.replace(day=1)
        default_end = today
        date_range = st.date_input("📅 Date Range", value=(default_start, default_end), disabled=ignore_date)

        if not ignore_date and all(date_range):
            st.caption(f"🗓 Selected range: {date_range[0].strftime('%d-%m-%Y')} to {date_range[1].strftime('%d-%m-%Y')}")

    logic = st.radio("Filter Logic", ["AND", "OR"], horizontal=True)

    filters = {
        "Shape": shape,
        "Color": color,
        "Quality": quality,
        "Seller": seller,
        "Buyer": buyer,
        "Pointer": pointer,
        "Size (mm)": size
    }

    # Parse 'Date' for filtering
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y", errors="coerce")

    if logic == "AND":
        for col, values in filters.items():
            if values and col in df.columns:
                df = df[df[col].isin(values)]
        if not ignore_date and all(date_range) and "Date" in df.columns:
            start_date, end_date = date_range
            df = df[(df["Date"] >= pd.to_datetime(start_date)) & (df["Date"] <= pd.to_datetime(end_date))]
    else:
        mask = pd.Series(False, index=df.index)
        for col, values in filters.items():
            if values and col in df.columns:
                mask |= df[col].isin(values)
        if not ignore_date and all(date_range) and "Date" in df.columns:
            mask |= (df["Date"] >= pd.to_datetime(date_range[0])) & (df["Date"] <= pd.to_datetime(date_range[1]))
        df = df[mask]

    # Format 'Date' back to dd-mm-yyyy for display
    if "Date" in df.columns:
        df["Date"] = df["Date"].dt.strftime("%d-%m-%Y").fillna("")

    return df

# Main logic
if uploaded_file:
    full_df = load_data(uploaded_file)
    if not full_df.empty:
        show_segmented_view(full_df)
else:
    st.info("📎 Please upload an Excel file to begin.")
       
        
      
    
      
    

   
